from langchain.text_splitter import RecursiveCharacterTextSplitter
from typing import List, Dict
import logging
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import pandas as pd
import io
import chardet

logger = logging.getLogger(__name__)

class DocumentProcessingError(Exception):
    """Exceção customizada para erro de processamento de documento."""
    pass

class DocumentProcessor:
    # Inicializa o splitter de texto para segmentar documento longo
    def __init__(self, chunk_size=1000, chunk_overlap=200):
        """
        Inicializa o RecursiveCharacterTextSplitter com o tamanho de chunk e sobreposição de chunk fornecidos.

        Parâmetros:
            chunk_size (int): O tamanho máximo de cada chunk. Padrão é 1000.
            chunk_overlap (int): O número de caracteres para sobrepor entre chunks. Padrão é 200.

        Retorna:
            None
        """
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap
        )

    def process_file(self, file_content: bytes, filename: str) -> List[Dict]:
        """
        Processa um arquivo e divide o seu conteúdo em segmentos menores.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo a ser processado.
            filename (str): O nome do arquivo, utilizado para determinar o tipo de arquivo e armazenar metadados.

        Retorna:
            List[Dict]: Uma lista de dicionários, onde cada dicionário contém o conteúdo de um segmento do arquivo e metadados sobre o arquivo original.
        """
        try:
            # Determina o tipo de arquivo e extrai o texto apropriadamente
            file_extension = filename.split('.')[-1].lower()
            if file_extension == 'pdf':
                text = self._extract_text_from_pdf(file_content)
            elif file_extension in ['doc','docx']:
                text = self._extract_text_from_docx(file_content)
            elif file_extension in ['xlsx', 'xls']:
                text = self._extract_text_from_excel(file_content)
            elif file_extension in ['htm', 'html']:
                text = self._extract_text_from_html(file_content)
            elif file_extension == 'csv':
                text = self._extract_text_from_csv(file_content)
            elif file_extension in ['txt', 'json', 'md']:
                text = self._extract_text_from_generic(file_content)
            else:
                raise DocumentProcessingError(f"Formato de arquivo não suportado: {file_extension}")

            # Divide o texto em segmentos menores
            segments = self.text_splitter.split_text(text)
            # Retorna o conteúdo de cada segmento e metadados sobre o arquivo original
            return [{"content": seg, "metadata": {"source": filename}} for seg in segments]
        except Exception as e:
            logger.error(f"Erro ao processar arquivo {filename}: {str(e)}")
            raise DocumentProcessingError(f"Falha ao processar arquivo {filename}: {str(e)}")

    def _extract_text_from_pdf(self, file_content: bytes) -> str:
        """
        Extrai texto de um arquivo PDF.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo PDF a ser processado.

        Retorna:
            str: O texto extraído do arquivo PDF.
        """
        pdf = PdfReader(io.BytesIO(file_content))
        return " ".join(page.extract_text() for page in pdf.pages)

    def _extract_text_from_docx(self, file_content: bytes) -> str:
        """
        Extrai texto de um arquivo Word.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo Word a ser processado.

        Retorna:
            str: O texto extraído do arquivo Word.
        """
        doc = DocxDocument(io.BytesIO(file_content))
        return " ".join(paragraph.text for paragraph in doc.paragraphs)

    def _extract_text_from_excel(self, file_content: bytes) -> str:
        """
        Extrai texto de um arquivo Excel.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo Excel a ser processado.

        Retorna:
            str: O texto extraído do arquivo Excel.
        """
        workbook = load_workbook(filename=io.BytesIO(file_content))
        text = []
        for sheet in workbook.sheetnames:
            for row in workbook[sheet].iter_rows(values_only=True):
                text.append(" ".join(str(cell) for cell in row if cell))
        return " ".join(text)

    def _extract_text_from_html(self, file_content: bytes) -> str:
        """
        Extrai texto de um arquivo HTML.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo HTML a ser processado.

        Retorna:
            str: O texto extraído do arquivo HTML.
        """
        soup = BeautifulSoup(file_content, 'html.parser')
        return soup.get_text()

    def _extract_text_from_csv(self, file_content: bytes) -> str:
        """
        Extrai texto de um arquivo CSV.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo CSV a ser processado.

        Retorna:
            str: O texto extraído do arquivo CSV.
        """
        df = pd.read_csv(io.BytesIO(file_content))
        return df.to_string(index=False)

    def _extract_text_from_generic(self, file_content: bytes) -> str:
        """
        Tenta detectar a codificação e extrair texto de um arquivo genérico.

        Parâmetros:
            file_content (bytes): O conteúdo do arquivo genérico a ser processado.

        Retorna:
            str: O texto extraído do arquivo genérico.
        """
        encoding = chardet.detect(file_content)['encoding']
        return file_content.decode(encoding)

    def process_multiple_documents(self, documents: List[Dict]) -> List[Dict]:
        """
        Processa documentos em lote.

        Parâmetros:
            documents (List[Dict]): Uma lista de dicionários, onde cada dicionário representa um documento.
                Cada dicionário deve conter as chaves 'content' e 'filename'.

        Retorna:
            List[Dict]: Uma lista de dicionários, onde cada dicionário representa um documento processado.
        """
        processed_documents = []
        for doc in documents:
            try:
                processed_documents.extend(self.process_file(doc['content'], doc['filename']))
            except DocumentProcessingError as e:
                logger.error(f"Erro ao processar documento: {str(e)}")
                
        return processed_documents